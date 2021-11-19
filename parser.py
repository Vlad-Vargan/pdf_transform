import re
from typing import Optional
from typing import List
from typing import Tuple

import pdftotext
from openpyxl import load_workbook


class ExcelConverter:

    init_row = 9
    col_range = ("B", "C", "D", "E")
    template_path = "template.xlsx"

    @classmethod
    def convert(cls, data: List[Tuple], filename: str) -> str:

        wb = load_workbook(filename=cls.template_path)
        # Pillow installation is necessary for saving imgs
        # grab the active worksheet
        cls.ws = wb.active

        for page, values in enumerate(data):
            rowx = str(cls.init_row + page)
            for colx, value in zip(cls.col_range, values):
                cls.ws[colx+rowx] = value

        # ws.move_range("A383:E397", rows=-1*page)
        table_end = len(data) + cls.init_row
        shift = -1 * (383 - table_end)
        cls._move_range(row_range=(383, 398), coll_range=(1, 6), row_shift=shift)

        cls._merge_table(start_row=table_end)

        cls.ws.delete_rows(table_end+14, amount=500)

        # Save the file
        save_path = filename.replace(".pdf", ".xlsx")
        wb.save("uploads/" + save_path)
        return save_path

    @classmethod
    def _merge_table(cls, start_row: int):
        ws = cls.ws
        for row in range(start_row+1, start_row+7):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        seven = start_row + 7
        ws.merge_cells(start_row=seven, start_column=1, end_row=seven, end_column=2)
        eight = start_row + 8
        ws.merge_cells(start_row=eight, start_column=1, end_row=eight, end_column=5)
        nine = start_row + 9
        ws.merge_cells(start_row=nine, start_column=1, end_row=nine, end_column=3)
        for row in range(start_row + 10, start_row + 16):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=start_row + 11, start_column=4, end_row=start_row + 13, end_column=5)
        ws[f"E{start_row}"] = f"=SUM(E{cls.init_row}:E{start_row-1})"
        ws[f"E{start_row+6}"] = f"=E{start_row}-SUM(E{start_row+1}:E{start_row+5})"

    @classmethod
    def _move_cell(cls, to_row: int, to_col: int, from_row: int, from_col: int):
        ws = cls.ws
        ws.cell(row=to_row, column=to_col).value = ws.cell(row=from_row, column=from_col).value
        from_style = ws.cell(row=from_row, column=from_col)._style
        ws.cell(row=to_row, column=to_col)._style = from_style
        ws.cell(row=to_row, column=to_col).hyperlink = ws.cell(row=from_row, column=from_col).hyperlink
        ws.cell(row=to_row, column=to_col).number_format = ws.cell(row=from_row, column=from_col).number_format

    @classmethod
    def _move_range(cls, row_range: Tuple, coll_range: Tuple, row_shift: int = 0, col_shift: int = 0):
        for from_row in range(row_range[0], row_range[1]):
            for from_col in range(coll_range[0], coll_range[1]):
                to_row = from_row + row_shift
                to_col = from_col + col_shift
                cls._move_cell(to_row, to_col, from_row, from_col)

class PDFParser:

    company_name = r"^((C\/O Orange Commercial Credit))$"

    invoice_number_pattern = r"^(Invoice\sNumber\s(\d{6}-\d-\w)\sPrint\sDate)$"
    ref_number_pattern = r"^(REF #1: ([a-zA-Z0-9-./ ]{1,25}))$"
    debtors_name_pattern = r"^(Bill To: ([-\w/() ]*))$"
    invoice_amount_pattern = r"^(Total Charges: \$ US\s*((\d{1,3}|\d{1},\d{3})\.\d{2}))$"

    @classmethod
    def parse(cls, filename: str) -> str:

        with open(filename, "rb") as f:
            pdf = pdftotext.PDF(f)

        data = []
        page, total_pages = 0, len(pdf)

        while page < total_pages:
            pdf_page = pdf[page]

            company_name = PDFParser._extract_pattern(pdf_page, cls.company_name)

            if not company_name:
                page += 1
                continue

            inv_no = PDFParser._extract_pattern(pdf_page, cls.invoice_number_pattern)
            po_no = PDFParser._extract_pattern(pdf_page, cls.ref_number_pattern)
            debtors_name = PDFParser._extract_pattern(pdf_page, cls.debtors_name_pattern)
            invoice_amount = PDFParser._extract_pattern(pdf_page, cls.invoice_amount_pattern)

            print(str(page+1) + ": ", company_name, "\t", inv_no, "\t", po_no, "\t", debtors_name, "\t", invoice_amount)

            if inv_no:
                if not invoice_amount:
                    page += 1
                    next_page = pdf[page]
                    invoice_amount = PDFParser._extract_pattern(next_page, cls.invoice_amount_pattern)
                invoice_amount = float(invoice_amount.replace(",", ""))
                values = (inv_no, po_no, debtors_name, invoice_amount)
                data.append(values)

            page += 1

        return data

    @staticmethod
    def _extract_pattern(pdf_page: str, pattern: str) -> str:
        if match := re.search(pattern, pdf_page, re.MULTILINE):
            return match.group(2)
        return ""


if __name__ == '__main__':
    data = PDFParser.parse("uploads/big_name.pdf")
    ExcelConverter.convert(data)